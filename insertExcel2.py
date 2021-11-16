# -*- coding: utf-8 -*-
"""
Spyder Editor

This is a temporary script file.
"""

import os
import json
import openpyxl
from openpyxl.styles import colors
from openpyxl.styles import Color
from openpyxl.styles import Border, Side
from openpyxl import load_workbook
import datetime
import sys


BORDER_THIN = 'thin'
BORDER_THICK = 'thick'



    
    
    
DIR = os.getenv("ATAM_HOME")

for files in os.listdir(DIR):
    #print(files)
    if 'ATAMContro' in files:
        #print(files)
        controller_directory = files;
        print(controller_directory);
        break;

config_path = DIR + "/" + controller_directory
f = open(config_path + "/config.properties", 'r')
line = f.readline()
print("-------------")
print(line)
line = line[9:]
line = line.split('=')[1]
line = line.replace(' ','')
line = line.replace(' ','')
line = line.replace('\n','')
line = line.replace('\n','')
line = line.replace('\n','')
dbhost = line
print("접속IP = " + dbhost)

###############################################################################
#
#
#                          변수 설정
#
###############################################################################


if (len(sys.argv) > 1):
    start_date = sys.argv[1]
    end_date = sys.argv[2]
    v1 = sys.argv[3]
    v2 = sys.argv[4]
    v3 = sys.argv[5]
    v4 = sys.argv[6]
    v5 = sys.argv[7]
    
    test_perform = "X"

    review_content = v2
    app_nm = v3
    app_version = v4
    reply_content = v5
    
    print("스토어등록일은 \n -" + v1)
    print("리뷰내용은 \n -" + review_content)
    print("서비스 종류는 \n -" + app_nm)
    print("앱버전은 \n -" + app_version)
    print("답변내용은 \n -" + reply_content)
    
#    v1 = v1.split("년")
#    v1_year = v1[0]
#    print(len(v1))
#    v1_year_after = v1[1].split("월")
#    v1_month = v1_year_after[0].replace(" ", "")
#    v1_day = v1_year_after[1].replace("일","").replace(" ", "")
#    if(int(v1_month) < 10):
#        v1_month = "0" + v1_month
#    if(int(v1_day) < 10):
#        v1_day = "0" + v1_day
#    
    v1_split = v1.split("-")
    v1_year = v1_split[0]
    v1_month = v1_split[1]
    v1_day = v1_split[2]
    if(int(v1_month) < 10):
        v1_month = "0" + v1_month
    if(int(v1_day) < 10):
        v1_day = "0" + v1_day
    
    v1_ymd = v1_year + v1_month + v1_day
    
    
    
    start_date = int(start_date)
    end_date = int(end_date)
    v1_ymd = int(v1_ymd)
    v1_cross_ymd = v1_year + "-" + v1_month + "-" + v1_day
    aa = datetime.datetime.strptime(v1_cross_ymd, "%Y-%m-%d").date()
    
    
    if (start_date <= v1_ymd <= end_date ):
        service_date = v1_year + "-" + v1_month + "-" + v1_day
    else:
        service_date = v1_year + "-" + v1_month + "-" + v1_day
        sys.exit(service_date + "는 날짜 기준과 맞지 않습니다.")
        
    
    
    
    if (len(reply_content) < 3):
        reply_category = "X"
    else :
        reply_category = "O"
    
    
    
    
else:
    start_date = ""
    end_date = ""
    v1 = ""
    v2 = ""
    v3 = ""
    v4 = ""
    v5 = ""
    test_perform = ""
    service_date = v1
    review_content = v2
    app_nm = v3
    app_version = v4
    reply_content = v5
    reply_category=""




try:
    r = open('c:/atam/insertExcelFile.txt', mode='r', encoding='utf-8')
    line = r.readline() #파일의 라인 끝에 줄 바꿈 (\n) 이 있을 경우 줄바꿈을 포함합니다.
    print("파일을 읽어왔습니다.")
    filename = line.strip()
    r.close()
except:
    r = open('c:/atam/insertExcelFile.txt', mode ='w', encoding='utf-8')
    r.write("c:/atam/template.xlsx")
    filename = "C:/atam/template.xlsx"
    print("파일명이 새로 저장되었습니다")
    r.close()



if filename != "c:/atam/template.xlsx" :
    excel_path = filename
    print("성공입니다")
else :
    excel_path = 'C:/atam/template.xlsx'
    print("실패입니다.")
excel_name = excel_path.split('/')
excel_name_lenth = len(excel_name)
file_nm = excel_name[excel_name_lenth - 1]
print(file_nm)
print(excel_name_lenth)
load_wb = load_workbook(excel_path, data_only=True)
sheet = load_wb.active




last_row = sheet.max_row
while sheet.cell(column=2, row=last_row).value is None and last_row > 0:
    last_row -= 1
last_col_a_value = sheet.cell(column=2, row=last_row).value
print("No는" + str(last_col_a_value))

print(last_row)


for sheet in load_wb.worksheets:
    for row in sheet.iter_rows():
        for entry in row:
            try:
                if 'No' in entry.value:
                    #print(entry.offset(row=0).value)

                    rows_cols = str(entry)
                    print("-------------------")
            except (AttributeError, TypeError) :
                  continue

for sheet in load_wb.worksheets:
    for row in sheet.iter_rows():
        for entry in row:
            try:
                if '비고' in entry.value:
                    #print(entry.offset(row=0).value)

                    rows_cols_end = str(entry)
                    print("-------------------")
            except (AttributeError, TypeError) :
                  continue



rows_cols = rows_cols.split('.')



rows_cols_name = rows_cols[len(rows_cols) -1].replace('>','')
print("rows_cols_name은 = " + rows_cols_name)

rows_cols_end = rows_cols_end.split('.')
rows_cols_end_name = rows_cols_end[len(rows_cols_end) -1].replace('>','')


print("현재 입력된 셀의 세로 범위는 = " + rows_cols_name[1:])
print("현재 입력된 셀의 가로 범위는 = " + str(sheet.max_column))

rows_cols_name_spelling = rows_cols_name[:1]
print("No가 들어갈 셀의 알파벳은 = " + rows_cols_name_spelling)
print("rows_cols_name_spelling은 = " + rows_cols_name_spelling)
last_row_start = rows_cols_name_spelling + str(last_row)
spellingtonumber = ord(rows_cols_name_spelling)
print("스펠링을 ord로 변경하면 = " + str(ord(rows_cols_name_spelling)))
rows_cols_start_num = int(rows_cols_name[1:]) + 1
rows_cols_start_num = last_row
print("입력시작열은" + str(rows_cols_start_num + 1))
print("max_rows는 " +str(sheet.max_row))


##############################################################################
#
#
#                       셀 설정
#
##############################################################################

cellalpharbet_0 = chr(spellingtonumber)
cellalpharbet_1 = chr(spellingtonumber + 1)
cellalpharbet_2 = chr(spellingtonumber + 2)
cellalpharbet_3 = chr(spellingtonumber + 3)
cellalpharbet_4 = chr(spellingtonumber + 4)
cellalpharbet_5 = chr(spellingtonumber + 5)
cellalpharbet_6 = chr(spellingtonumber + 6)
cellalpharbet_7 = chr(spellingtonumber + 7)
cellalpharbet_8 = chr(spellingtonumber + 8)
cellalpharbet_9 = chr(spellingtonumber + 9)
cellalpharbet_10 = chr(spellingtonumber + 10)
cellalpharbet_11 = chr(spellingtonumber + 11)
cellalpharbet_12 = chr(spellingtonumber + 12)
cellalpharbet_13 = chr(spellingtonumber + 13)
cellalpharbet_14 = chr(spellingtonumber + 14)
cellalpharbet_15 = chr(spellingtonumber + 15)
cellalpharbet_16 = chr(spellingtonumber + 16)
cellalpharbet_17 = chr(spellingtonumber + 17)
cellalpharbet_18 = chr(spellingtonumber + 18)
cellalpharbet_19 = chr(spellingtonumber + 19)
cellalpharbet_20 = chr(spellingtonumber + 20)
    #sheet.max_column = sheet.max_column + 1
    
cell_0 = cellalpharbet_0 + str(rows_cols_start_num + 1)
cell_1 = cellalpharbet_1 + str(rows_cols_start_num + 1)
cell_2 = cellalpharbet_2 + str(rows_cols_start_num + 1)
cell_3 = cellalpharbet_3 + str(rows_cols_start_num + 1)
cell_4 = cellalpharbet_4 + str(rows_cols_start_num + 1)
cell_5 = cellalpharbet_5 + str(rows_cols_start_num + 1)
cell_6 = cellalpharbet_6 + str(rows_cols_start_num + 1)
cell_7 = cellalpharbet_7 + str(rows_cols_start_num + 1)
cell_8 = cellalpharbet_8 + str(rows_cols_start_num + 1)
cell_9 = cellalpharbet_9 + str(rows_cols_start_num + 1)
cell_10 = cellalpharbet_10 + str(rows_cols_start_num + 1)
cell_11 = cellalpharbet_11 + str(rows_cols_start_num + 1)
cell_12 = cellalpharbet_12 + str(rows_cols_start_num + 1)
cell_13 = cellalpharbet_13 + str(rows_cols_start_num + 1)
cell_14 = cellalpharbet_14 + str(rows_cols_start_num + 1)
cell_15 = cellalpharbet_15 + str(rows_cols_start_num + 1)
cell_16 = cellalpharbet_16 + str(rows_cols_start_num + 1)
cell_17 = cellalpharbet_17 + str(rows_cols_start_num + 1)
cell_18 = cellalpharbet_18 + str(rows_cols_start_num + 1)
cell_19 = cellalpharbet_19 + str(rows_cols_start_num + 1)
cell_20 = cellalpharbet_20 + str(rows_cols_start_num + 1)
    
    
sheet[cell_0].border = Border(left  =Side(border_style=BORDER_THIN,
                                        color='000000'),
                              right =Side(border_style=BORDER_THIN,
                                         color='000000'),
                              top   =Side(border_style=BORDER_THIN,
                                       color='000000'),
                              bottom=Side(border_style=BORDER_THIN,
                                          color='000000'))
sheet[cell_1].border = Border(left=Side(border_style=BORDER_THIN,
                                        color='000000'),
                              right=Side(border_style=BORDER_THIN,
                                         color='000000'),
                              top=Side(border_style=BORDER_THIN,
                                       color='000000'),
                              bottom=Side(border_style=BORDER_THIN,
                                          color='000000'))
sheet[cell_2].border = Border(left=Side(border_style=BORDER_THIN,
                                        color='000000'),
                              right=Side(border_style=BORDER_THIN,
                                         color='000000'),
                              top=Side(border_style=BORDER_THIN,
                                       color='000000'),
                              bottom=Side(border_style=BORDER_THIN,
                                          color='000000'))    
sheet[cell_3].border = Border(left=Side(border_style=BORDER_THIN,
                                        color='000000'),
                              right=Side(border_style=BORDER_THIN,
                                         color='000000'),
                              top=Side(border_style=BORDER_THIN,
                                       color='000000'),
                              bottom=Side(border_style=BORDER_THIN,
                                          color='000000'))    
sheet[cell_4].border = Border(left=Side(border_style=BORDER_THIN,
                                        color='000000'),
                              right=Side(border_style=BORDER_THIN,
                                         color='000000'),
                              top=Side(border_style=BORDER_THIN,
                                       color='000000'),
                              bottom=Side(border_style=BORDER_THIN,
                                          color='000000'))    
sheet[cell_5].border = Border(left=Side(border_style=BORDER_THIN,
                                        color='000000'),
                              right=Side(border_style=BORDER_THIN,
                                         color='000000'),
                              top=Side(border_style=BORDER_THIN,
                                       color='000000'),
                              bottom=Side(border_style=BORDER_THIN,
                                          color='000000'))    
sheet[cell_6].border = Border(left=Side(border_style=BORDER_THIN,
                                        color='000000'),
                              right=Side(border_style=BORDER_THIN,
                                         color='000000'),
                              top=Side(border_style=BORDER_THIN,
                                       color='000000'),
                              bottom=Side(border_style=BORDER_THIN,
                                          color='000000'))
sheet[cell_7].border = Border(left=Side(border_style=BORDER_THIN,
                                        color='000000'),
                              right=Side(border_style=BORDER_THIN,
                                         color='000000'),
                              top=Side(border_style=BORDER_THIN,
                                       color='000000'),
                              bottom=Side(border_style=BORDER_THIN,
                                          color='000000'))    
sheet[cell_8].border = Border(left=Side(border_style=BORDER_THIN,
                                        color='000000'),
                              right=Side(border_style=BORDER_THIN,
                                         color='000000'),
                              top=Side(border_style=BORDER_THIN,
                                       color='000000'),
                              bottom=Side(border_style=BORDER_THIN,
                                          color='000000'))
sheet[cell_9].border = Border(left=Side(border_style=BORDER_THIN,
                                        color='000000'),
                              right=Side(border_style=BORDER_THIN,
                                         color='000000'),
                              top=Side(border_style=BORDER_THIN,
                                       color='000000'),
                              bottom=Side(border_style=BORDER_THIN,
                                          color='000000'))
sheet[cell_10].border = Border(left=Side(border_style=BORDER_THIN,
                                        color='000000'),
                              right=Side(border_style=BORDER_THIN,
                                         color='000000'),
                              top=Side(border_style=BORDER_THIN,
                                       color='000000'),
                              bottom=Side(border_style=BORDER_THIN,
                                          color='000000'))
sheet[cell_11].border = Border(left=Side(border_style=BORDER_THIN,
                                        color='000000'),
                              right=Side(border_style=BORDER_THIN,
                                         color='000000'),
                              top=Side(border_style=BORDER_THIN,
                                       color='000000'),
                              bottom=Side(border_style=BORDER_THIN,
                                          color='000000'))
sheet[cell_12].border = Border(left=Side(border_style=BORDER_THIN,
                                        color='000000'),
                              right=Side(border_style=BORDER_THIN,
                                         color='000000'),
                              top=Side(border_style=BORDER_THIN,
                                       color='000000'),
                              bottom=Side(border_style=BORDER_THIN,
                                          color='000000'))
sheet[cell_13].border = Border(left=Side(border_style=BORDER_THIN,
                                        color='000000'),
                              right=Side(border_style=BORDER_THIN,
                                         color='000000'),
                              top=Side(border_style=BORDER_THIN,
                                       color='000000'),
                              bottom=Side(border_style=BORDER_THIN,
                                          color='000000'))
sheet[cell_14].border = Border(left=Side(border_style=BORDER_THIN,
                                        color='000000'),
                              right=Side(border_style=BORDER_THIN,
                                         color='000000'),
                              top=Side(border_style=BORDER_THIN,
                                       color='000000'),
                              bottom=Side(border_style=BORDER_THIN,
                                          color='000000'))
sheet[cell_15].border = Border(left=Side(border_style=BORDER_THIN,
                                        color='000000'),
                              right=Side(border_style=BORDER_THIN,
                                         color='000000'),
                              top=Side(border_style=BORDER_THIN,
                                       color='000000'),
                              bottom=Side(border_style=BORDER_THIN,
                                          color='000000'))
sheet[cell_16].border = Border(left=Side(border_style=BORDER_THIN,
                                        color='000000'),
                              right=Side(border_style=BORDER_THIN,
                                         color='000000'),
                              top=Side(border_style=BORDER_THIN,
                                       color='000000'),
                              bottom=Side(border_style=BORDER_THIN,
                                          color='000000'))
sheet[cell_17].border = Border(left=Side(border_style=BORDER_THIN,
                                        color='000000'),
                              right=Side(border_style=BORDER_THIN,
                                         color='000000'),
                              top=Side(border_style=BORDER_THIN,
                                       color='000000'),
                              bottom=Side(border_style=BORDER_THIN,
                                          color='000000'))
sheet[cell_18].border = Border(left=Side(border_style=BORDER_THIN,
                                        color='000000'),
                              right=Side(border_style=BORDER_THIN,
                                         color='000000'),
                              top=Side(border_style=BORDER_THIN,
                                       color='000000'),
                              bottom=Side(border_style=BORDER_THIN,
                                          color='000000'))
sheet[cell_19].border = Border(left=Side(border_style=BORDER_THIN,
                                        color='000000'),
                              right=Side(border_style=BORDER_THIN,
                                         color='000000'),
                              top=Side(border_style=BORDER_THIN,
                                       color='000000'),
                              bottom=Side(border_style=BORDER_THIN,
                                          color='000000'))
sheet[cell_20].border = Border(left=Side(border_style=BORDER_THIN,
                                        color='000000'),
                              right=Side(border_style=BORDER_THIN,
                                         color='000000'),
                              top=Side(border_style=BORDER_THIN,
                                       color='000000'),
                              bottom=Side(border_style=BORDER_THIN,
                                          color='000000'))


#########################################################################################
#
#
#                      엑셀 입력
#
#########################################################################################




#sheet[cellzero]last_col_a_value
if last_col_a_value == 'No':
    last_col_a_value = '0'
sheet[cell_0] = str(int(last_col_a_value)+1)
sheet[cell_0].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

sheet[cell_1] = str(service_date)
sheet[cell_1].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

sheet[cell_2] = str(review_content)
sheet[cell_2].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

sheet[cell_3] = str(app_nm)
sheet[cell_3].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

sheet[cell_4] = '기타'
sheet[cell_4].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

sheet[cell_5] = str(app_version)
sheet[cell_5].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

sheet[cell_6] = str(reply_category)
sheet[cell_6].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

sheet[cell_7] = str(reply_content)
sheet[cell_7].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

sheet[cell_8] = test_perform
sheet[cell_8].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')




dt_now = datetime.datetime.now()
dt_now = str(dt_now).split('.')
#
dt_now = dt_now[0]
dt_now = dt_now.replace('-','').replace(':','').replace('.','').replace(' ','_')




if (filename == "C:/atam/template.xlsx"):
    r = open('c:/atam/insertExcelFile.txt', mode ='w', encoding='utf-8')
    filename = "C:/atam/" +dt_now+ ".xlsx"
    print(filename)
    r.write(filename)
    print("파일명이 새로 저장되었습니다")
    r.close()



print("저장 명은 " + filename + "입니다")
load_wb.save(filename)






